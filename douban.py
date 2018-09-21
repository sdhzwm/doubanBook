import time
import random
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

def bookTag():

    response = requests.get('https://book.douban.com/tag/?view=cloud')

    soup = BeautifulSoup(response.text, 'lxml')

    tags = soup.select('.tagCol td a')
    tag_list = []
    start = 0
    end = 7
    for tag in tags:
        tag_list.append(tag.get_text())
    tr, td = divmod(len(tag_list), 7)
    # print(tr, td)
    if td != 0:
        tr = tr + 1
    for i in range(tr):
        # print(tag_list[start:end])
        start, end = end, end + 7
    return tag_list

def devideTag(book_tag_list):
    # 操作全局变量
    global START_PAGE
    book_list = []
    for book_tag in book_tag_list:
        # 每个标签下的书籍
        book_list = bookSpider(book_tag)
        # 每爬完一个标签，将起始页归为1
        START_PAGE = 1
        end = 1
    print_book_lists_excel(book_list,book_tag_list)



def bookSpider(book_tag):
    global START_PAGE
    end = 1
    book_list = []
    # 设置请求头
    hds = [
            {'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
            {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
            {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}
           ]
    # headers = {
    #     'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36'}

    while START_PAGE <= int(end):
        param = 'start=' + str((START_PAGE - 1) * 20) + '&type=T'
        url = url = 'https://book.douban.com/tag/%s/' % book_tag
        time.sleep(random.random() * 2)
        # 如果网址有错误，报错、退出
        try:
            # 拼接url
            response = requests.get(url, params=param, headers=hds[end%len(hds)])

        except Exception as e:
            print(e)
            break
        # 解析页面
        soup = BeautifulSoup(response.text, 'lxml')
        # 这个页面是否有内容，如果没有退出
        try:
            content = soup.find_all('li', class_='subject-item')
            if len(content) == 0:
                break
        except:

            break
        # 获取你想要的数据

        b_list = loadData(content, book_tag,book_list)
        end += 1
    return b_list


def loadData(content, book_tag,book_list):
    global START_PAGE
    for book_info in content:
        # 书名需要处理一下，split以空格分隔，移除空字符串，返回一个列表
        f_title = book_info.select('.info h2')[0].get_text().split()
        # 连接字符串
        b_title = ''.join(f_title)
        pub_info= book_info.select('.info .pub')[0].get_text().strip()
        desc_list = pub_info.split('/')
        book_url = '图书链接地址' + book_info.select('.info h2 a')[0].get('href')
        pic_url = '图片链接地址' + book_info.select('.pic img')[0].get('src')
        try:
            book_des = book_info.select('.info p')[0].get_text()
        except:
            book_des = '暂无详情'
        try:
            author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
        except:
            author_info = '作者/译者： 暂无'
        try:
            pub_info = '出版信息： ' + '/'.join(desc_list[-3:-1])
        except:
            pub_info = '出版信息： 暂无'
        try:
            rating = book_info.select('.info .rating_nums')[0].get_text().strip()
        except:
            rating = '0.0'

        book_list.append([b_title, rating, book_des,author_info, pub_info,book_url,pic_url])
    print('下载页面的内容：' % START_PAGE,book_tag)
    # 页数加1
    START_PAGE += 1
    return book_list

def print_book_lists_excel(book_lists,book_tag_lists):
    wb = Workbook()
    ws=[]
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i]))
    for i in range(len(book_tag_lists)):
        ws[i].append(['序号','书名','评分','简介','作者','出版信息','大图','小图'])
        count=1
        for bl in book_lists:
            ws[i].append([count,bl[0],bl[1],bl[2],bl[3],bl[4],bl[5],bl[6]])
            count+=1
    save_path='book'
    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i])
    save_path+='.xlsx'
    wb.save(save_path)



if __name__ == '__main__':
    START_PAGE = 1
    book_tag_list = []
    book_lists = []
    tag_list = []
    tag_list = bookTag()

    for book_tag in tag_list:
        book_tag_list.append(book_tag)
        devideTag(book_tag_list)

